# my_excel_lib/validation.py

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

def apply_validation(final_output_file_path):
    try:
        wb = load_workbook(final_output_file_path)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active
    ws.freeze_panes = 'A2'  # Freeze top row
    print("Top row Freezed")

    choices = {
        "Result": ['Fixed', 'LE', 'Blocked', 'Missing', 'Pure Edit', 'Not found'],
        "Result (validation)": ['Fixed', 'LE', 'Blocked', 'Missing', 'Pure Edit', 'Not found'],
        "Status": ['RFQC', 'Spec 1.1', 'Already Pass', 'Blocker', 'Already Blocked', 'Repeated', 'Not found'],
        "Error Found?": ['Yes', 'No'],
        "LBC Alignment": ['Yes', 'No'],
        "Observed_Lane_Type": ["Regular", "Shoulder", "Acceleration", "Deceleration", "Auxiliary", "Drivable shoulder", "Bicycle", "Other", "Hov", "Reversible", "Slow", "Express", "Center turn", "Turn", "Drivable parking", "On street parking", "Variable driving"],
        "Updated_Lane_Type": ["Regular", "Shoulder", "Acceleration", "Deceleration", "Auxiliary", "Drivable shoulder", "Bicycle", "Other", "Hov", "Reversible", "Slow", "Express", "Center turn", "Turn", "Drivable parking", "On street parking", "Variable driving"],
        "Smoothening Successful? ( LCL )": ['Yes', 'No'],
        "Validation Marked as LE? ( LCL )": ['Yes', 'No'],
        "Did Smoothening work? ( LCL )": ['Yes', 'No'],
        "Smoothening current status? ( LCL )": ['Default', 'Smoothed offset v1', 'Smoothed offset v2','Both didn\'t work'],
        "QC1 Result (Validation)": ['Fixed', 'LE', 'Blocked', 'Missing', 'Pure Edit', 'Not found'],
        "QC1 Status": ['RFQC', 'Spec 1.1', 'Already Pass', 'Blocker', 'Already Blocked', 'Repeated', 'Not found'],
        "LCL Smoothened? ( LCL )": ['Yes', 'No'],
        "QC2 Result (Validation)": ['Fixed', 'LE', 'Blocked', 'Missing', 'Pure Edit', 'Not found'],
        "QC2 Status": ['RFQC', 'Spec 1.1', 'Already Pass', 'Blocker', 'Already Blocked', 'Repeated', 'Not found'],
        "QC3 Result (Validation)": ['Fixed', 'LE', 'Blocked', 'Missing', 'Pure Edit', 'Not found'],
        "QC3 Status": ['RFQC', 'Spec 1.1', 'Already Pass', 'Blocker', 'Already Blocked', 'Repeated', 'Not found']
    }

    for column_name, column_choices in choices.items():
        column_index = None
        for cell in ws[1]:  # Assuming headers are in the first row
            if cell.value == column_name:
                column_index = cell.column
                break

        if column_index is not None:
            for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index, max_row=ws.max_row):
                for cell in row:
                    dv = DataValidation(type="list", formula1='"' + ','.join(column_choices) + '"', allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(cell)

    wb.save(final_output_file_path)
