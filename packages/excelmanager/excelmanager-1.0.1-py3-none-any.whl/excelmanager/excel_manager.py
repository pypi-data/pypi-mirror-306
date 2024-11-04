import os
from collections import OrderedDict
from typing import Any, Literal, Sequence, overload

import openpyxl
import openpyxl.cell
import openpyxl.styles


@overload
def read_file_sheet(
    input_file_path: str,
    sheet_name: str,
    output_row_type: Literal["OrderedDict"] = ...,
    has_headers: bool = ...,
    range: tuple[str, str] | None = ...,
) -> list[OrderedDict[Any, Any]]: ...


@overload
def read_file_sheet(
    input_file_path: str,
    sheet_name: str,
    output_row_type: Literal["OrderedDict"],
    has_headers: bool = ...,
    range: tuple[str, str] | None = ...,
) -> list[OrderedDict[Any, Any]]: ...


@overload
def read_file_sheet(
    input_file_path: str,
    sheet_name: str,
    output_row_type: Literal["list"],
    has_headers: bool = ...,
    range: tuple[str, str] | None = ...,
) -> list[list[Any]]: ...


@overload
def read_file_sheet(
    input_file_path: str,
    sheet_name: str,
    output_row_type: Literal["tuple"],
    has_headers: bool = ...,
    range: tuple[str, str] | None = ...,
) -> list[tuple[Any, ...]]: ...


def read_file_sheet(
    input_file_path: str,
    sheet_name: str,
    output_row_type: Literal["OrderedDict", "list", "tuple"] = "OrderedDict",
    has_headers=True,
    range: tuple[str, str] | None = None,
):
    """Reads data from an Excel file sheet.

    Args:
        input_file_path (str): The file path of the Excel file.
        sheet_name (str): The sheet name of the Excel file.
        output_row_type (Literal["OrderedDict", "list", "tuple"], optional): The row data type to use for returned data.
            Defaults to "OrderedDict".
        has_headers (bool, optional): Flag to determine whether the first row of the data should be used as headers when
            outputting rows in `OrderedDict` format. Defaults to True.
        range (tuple[str, str] | None, optional): A tuple specifying the cell range to read on the sheet,
            i.e. `("A1", "C10")`. Defaults to None.

    Returns:
        list[OrderedDict[Any, Any]] | list[list[Any]] | list[tuple[Any, ...]]: Data read from the Excel file.
    """

    sheet = openpyxl.load_workbook(input_file_path, data_only=True)[sheet_name]

    if range:
        range_str = f"{range[0]}:{range[1]}"
    else:
        range_str = sheet.dimensions

    data = sheet[range_str]

    headers = []
    output = []

    for i, row in enumerate(data):
        if has_headers and i == 0:
            headers = [cell.value for cell in row]

            if output_row_type == "list":
                output.append(headers)
            elif output_row_type == "tuple":
                output.append(tuple(headers))

            continue

        if output_row_type == "OrderedDict":
            output.append(_build_ordered_dict_from_row(headers, row))
        elif output_row_type == "list":
            output.append(_build_list_from_row(row))
        else:
            output.append(_build_tuple_from_row(row))

    return output


def write_file_sheet(
    output_file_name: str,
    output_file_directory: str,
    data: Sequence[OrderedDict[Any, Any]] | Sequence[Sequence[Any]],
    sheet_name="Sheet",
    headers: Sequence[str] | None = None,
    bold_headers=True,
):
    """Write data to an Excel file.

    Args:
        output_file_name (str): The file name (including the file extension) to write to or create.
        output_file_directory (str): The directory where the file is or will be located.
        data (Sequence[OrderedDict[Any, Any]] | Sequence[Sequence[Any]]): The data to write.
        sheet_name (str, optional): The Excel sheet to write to. Defaults to "Sheet1".
        headers (Sequence[str] | None, optional): A sequence of headers for the data. When the data row type is of
            `OrderedDict`, the dictionary keys of the first row will be used as headers unless an empty sequence is passed.
            Defaults to None.
        bold_headers (bool, optional): A flag to write the headers in bold font. Defaults to True.
    """
    file_path = os.path.join(output_file_directory, output_file_name)

    if os.path.isfile(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    sheet = workbook[sheet_name]
    dict_data = tuple(row for row in data if isinstance(row, OrderedDict))
    sequence_data = tuple(row for row in data if isinstance(row, Sequence))

    if headers is None and len(dict_data) > 0:
        headers = tuple(key for key in dict_data[0].keys())

    if dict_data:
        sequence_data = tuple(tuple(row.values()) for row in dict_data)

    bold_font = openpyxl.styles.Font(bold=True)

    if headers:
        for i, header in enumerate(headers):
            cell = sheet.cell(row=1, column=i + 1)
            cell.value = header

            if bold_headers:
                cell.font = bold_font

    for row in sequence_data:
        sheet.append(row)

    workbook.save(file_path)


def _build_ordered_dict_from_row(
    headers: Sequence[Any], row: Sequence[openpyxl.cell.Cell]
):
    has_headers = len(headers) > 0

    if has_headers and len(headers) != len(row):
        raise RuntimeError(
            f"Length {len(headers)} of headers does not match length {len(row)} of row."
        )

    dict_row: OrderedDict[Any, Any] = OrderedDict()

    for i, cell in enumerate(row):
        if has_headers:
            dict_row[headers[i]] = cell.value
        else:
            dict_row[i + 1] = cell.value

    return dict_row


def _build_list_from_row(row: Sequence[openpyxl.cell.Cell]):
    return [cell.value for cell in row]


def _build_tuple_from_row(row: Sequence[openpyxl.cell.Cell]):
    return tuple(cell.value for cell in row)
