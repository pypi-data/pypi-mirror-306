# -*- coding:UTF-8 -*-
"""
__title__ = '对Excel单元格的读写'
#
      1.pip install pandas
      2.pip install loguru
      3.pip install openpyxl
"""
import os
import traceback

from loguru import logger
from openpyxl.reader.excel import load_workbook
import pandas as pd


class ExcelUtil(object):
    """ excel 帮助类 """

    def __init__(self, xlsxPath=None, sheetname=None):
        """

        :param xlsxPath:
        :param sheetname: sheetname若为None 则默认读取索引为0即第一个sheet的内容
        """

        self._xlsxPath = xlsxPath
        self._sheetname = sheetname

    def getWbBook(self, sheet=None, header=None):
        """

        :param sheet:
        :param header: 源码里header默认是0,如果文件存在header不用传此参数,
        若文件无header,传None即可
        :return:
        """

        if sheet is not None and header is not None:
            self.df = pd.read_excel(self._xlsxPath, keep_default_na=False, dtype=object, sheet_name=sheet, header=header)
        elif sheet is not None and header is None:
            self.df = pd.read_excel(self._xlsxPath, keep_default_na=False, dtype=object, sheet_name=sheet)
        elif sheet is None and header is None:
            self.df = pd.read_excel(self._xlsxPath, keep_default_na=False, dtype=object)
        elif sheet is None and header is not None:
            self.df = pd.read_excel(self._xlsxPath, keep_default_na=False, dtype=object, header=header)
        else:
            self.df = pd.read_excel(self._xlsxPath, keep_default_na=False, dtype=object)
        return self.df

    def readHead(self):
        """
        读取标题行
        :return:
        """
        self.read_excel_to_dict()
        data = self.df.columns.values
        return data

    def readRow(self, rowNum):
        """
        读取某一行
        :param rowNum:
        :return:
        """
        self.read_excel_to_dict()
        data = self.df.iloc[rowNum].values
        return data

    def read_excel_to_dict(self):
        dataframe = pd.read_excel(self._xlsxPath, sheet_name=self._sheetname)
        if not self._sheetname:
            self.df = dataframe[list(dataframe.items())[0][0]]
            result = dataframe[list(dataframe.items())[0][0]].to_dict(orient='records')
        else:
            self.df = dataframe[self._sheetname]
            result = dataframe[self._sheetname].to_dict(orient='records')
        return result

    def dict_to_excel(self, data, excel_path, set_index=None):
        """  dict 转 excel
        :param data: {"ID":[1,2,3], "Name":["Tim", "ZhangSan", "LiSi"]}
        :param excel_path: "/Users/Cat/xxx.xls(x)"  # xlsx,xlx都支
        :param set_index: "ID"  # 将ID作为索引列
        :return:
        """
        df = pd.DataFrame(data)
        if set_index:
            df.set_index(set_index)
        df.to_excel(excel_path)

    def write_cell(self, row, col, value):
        """
        修改或新增单元格的值
        :param row: 要修改的单元格的行号 从0开始
        :param col: 要修改的单元格列号 从0开始
        :param value: 要写进单元格的值
        :return:
        """
        try:
            if self._xlsxPath is None:
                return logger.error("路径不能为None")
            else:
                if not os.path.exists(self._xlsxPath):
                    return logger.error(f"{self._xlsxPath} 文件不存在")
            if self._sheetname is None:
                return logger.error("sheetName不能为None")

            excl = load_workbook(self._xlsxPath)
            sheet = excl[self._sheetname]
            sheet.cell(row + 1, col + 1, value)
            excl.save(self._xlsxPath)
            return "SUCCESS"

        except Exception as ex:
            if "[Errno 13] Permission denied" in str(ex):
                return logger.error(f"{self._xlsxPath} 文件被占用,请先关闭文件")
            else:
                logger.error("\tError %s\n" % ex)
                return logger.error(traceback.format_exc())


if __name__ == '__main__':
    excel_util = ExcelUtil("7777777.xls")
    res = excel_util.read_excel_to_dict()
    # res_dict = excel_util.read_excel_to_dict()
    res_dict_list = excel_util.read_excel_to_dict()
    debug = True
    # 结果二维字典
    # data = excel_util.readExcel()
    # print(data[1][1])
    # 修改以及添加单元格的值
    # excel_util.write_cell(47, 1, "三翻四复")
    # excel_util.write_cell(47, 2, "快接啊号地块计划")
    # excel_util.write_cell(47, 3, "好人")
    # excel_util.write_cell(48, 1, "搜索vs")

    # xlsxPath = "D:\\2020交易网关\\TDGW配置模板.xls"
    # # 初始化对象
    # excel_util = ExcelUtil(xlsxPath)
    # # 初始化workbook对象
    # excel_util.getWbBook(0)
    # # 读取数据至dict
    # data = excel_util.read_to_dict()
    # print(data)
    #
    # for map in data:
    #     print(map['操作系统'])
