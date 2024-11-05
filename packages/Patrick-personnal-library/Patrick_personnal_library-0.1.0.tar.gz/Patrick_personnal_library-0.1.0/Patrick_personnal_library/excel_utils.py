import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os


def adjust_column_width(sheet):
    """Ajuster la largeur des colonnes en fonction du contenu."""
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Récupère la lettre de la colonne
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Ajoute une marge
        sheet.column_dimensions[col_letter].width = adjusted_width


def format_sheet(sheet):
    """Applique le formatage de style à une feuille Excel."""
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bold_font = Font(bold=True)

    for cell in sheet[1]:  # La première ligne est l'en-tête
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Appliquer des bordures fines
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def save_to_excel(data, to_add_for_file_name, headers, save_path, 
                  sheet_title="Sheet1", freeze_top=True, freeze_both=False, apply_formatting=True):
    """
    Enregistre les données dans un fichier Excel avec des options pour figer la première ligne 
    et appliquer le formatage.

    Args:
        data (list of list/tuple): Les données à ajouter à la feuille.
        to_add_for_file_name (str): Texte à ajouter dans le nom de fichier.
        headers (list): Les en-têtes de colonne.
        save_path (str): Chemin de sauvegarde du fichier.
        sheet_title (str, optional): Le titre de la feuille. Par défaut "Sheet1".
        freeze_top (bool, optional): Figer la première ligne si True. Par défaut True.
        apply_formatting (bool, optional): Appliquer le formatage si True. Par défaut True.
    """

    
    timestamp = datetime.now().strftime('%Y-%m-%d')
    file_name = f"{to_add_for_file_name}_{timestamp}.xlsx"
    file_path = os.path.join(save_path, file_name)

    # Charger ou créer un fichier Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    # Ajouter les en-têtes
    sheet.append(headers)

    # Write data rows
    for row in data:
        flattened_row = []
        for item in row:
            if isinstance(item, list):
                # Flatten the list and add each element to the row
                for sub_item in item:
                    if isinstance(sub_item, list):
                        # If sub_item is a list (like ['Tenacity', 19.6]), add each element separately
                        flattened_row.extend(sub_item)
                    else:
                        flattened_row.append(sub_item)
            else:
                flattened_row.append(item)
        # Append the flattened row to the sheet
        sheet.append(flattened_row)
        
    # Figer la première ligne et la première colonne si l'option est activée
    if freeze_both:
        sheet.freeze_panes = "B2"

    # Figer la première ligne si l'option est activée
    if freeze_top:
        sheet.freeze_panes = "A2"

    # Appliquer le formatage si activé
    if apply_formatting:
        format_sheet(sheet)
        adjust_column_width(sheet)

    # Sauvegarder le fichier Excel
    workbook.save(file_path)
    print(f"Données enregistrées dans {file_path}.")


def lire_cellule_excel(fichier_excel, feuille, cellule):
    try:
        # Charger le classeur
        classeur = load_workbook(fichier_excel, data_only=True)
        # Selectionner la feuille
        feuille = classeur[feuille]
        # Lire la valeur de la cellule
        valeur = feuille[cellule].value
        return valeur
    except Exception as e:
        print(f"Erreur lors de la lecture de la cellule {cellule} : {e}")

def convert_dict_to_list(data:dict):
    data_to_save = []
    for value in data.values():
        data_for_each_toon = []
        for key in value:
            data_for_each_toon.append(value[key])
        data_to_save.append(data_for_each_toon)
    return data_to_save

def construire_dict_cellule_colonne_for_DB(fichier_excel, feuille_name):
    try:
        classeur = load_workbook(fichier_excel, data_only=True)
        feuille = classeur[feuille_name]
        cellule_colonnes = {}
        for row in feuille.iter_rows(min_row=2, max_col=2, values_only=True):
            cellule, colonne = row
            if cellule and colonne:
                cellule_colonnes[cellule] = colonne
        return cellule_colonnes

    except Exception as e:
        print(f"Erreur lors de la construction du dictionnaire cellulkes_colonnes : {e}")
        return {}
    
def lire_une_cellule(fichier_excel, feuille_name, cellule_a_lire):
    try:
        # Charger le feuille
        classeur = load_workbook(fichier_excel, data_only=True)
        # Selectionner la feuille
        feuille = classeur[feuille]
        # Lire la valeur de la cellule
        valeur = feuille[cellule_a_lire].value
        return valeur
    except Exception as e:
        print(f"Erreur lors de la lecture de la cellule {cellule_a_lire} : {e}")

def formater_valeur(valeur, precision):
    try:
        return f"{valeur:.{precision}f}"
    except (ValueError, TypeError):
        return valeur
